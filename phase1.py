#!/usr/bin/env python3
"""
phase1.py — end-to-end infoflora photo pipeline (merged single-file version).

Merges: pipeline.py, download_photos.py, download_zeigerwerte.py,
        proto_geojson.py, genera_fioritura.py, genera_iucn.py,
        genera_zeigerwerte.py, chart_style.py

Usage: venv/bin/python3 phase1.py species.xlsx
"""

import sys
import os
import re
import io
import csv
import time
import shutil
import json
import base64
from datetime import datetime
from urllib.request import urlopen
from urllib.parse import urlparse

import requests
import openpyxl
from openpyxl.styles import Font
import numpy as np
import cv2
from scipy.ndimage import uniform_filter1d, gaussian_filter
from bs4 import BeautifulSoup
from PIL import Image, ImageDraw, ImageFont
import cairosvg


# ── CONFIG ───────────────────────────────────────────────────────────────────

OUTPUT_DIR         = "output"

MAIN_OUTPUT_DIR    = os.path.join(OUTPUT_DIR, "images_ext")
MAIN_COPYRIGHT     = os.path.join(OUTPUT_DIR, "_copyright_main.csv")
MAIN_CANDIDATES    = os.path.join(OUTPUT_DIR, "images_ext_candidates")

JOINED_XLSX        = os.path.join(OUTPUT_DIR, "_joined_intermediate.xlsx")

SPECIAL_INPUT_CSV  = os.path.join(OUTPUT_DIR, "_special_cases_input.csv")
SPECIAL_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "special_cases")
SPECIAL_COPYRIGHT  = os.path.join(OUTPUT_DIR, "_copyright_special.csv")
SPECIAL_CANDIDATES = os.path.join(OUTPUT_DIR, "special_cases_candidates")

LOG_PATH           = os.path.join(OUTPUT_DIR, "pipeline.log")

MAPS_DIR           = os.path.join(OUTPUT_DIR, "maps")

CSV_EXTRA_COLS     = ["File", "Credit", "Status", "Map"]
HEADER_KEYWORDS    = {"taxonname", "nom du taxon"}

INTERMEDIATES      = [MAIN_COPYRIGHT, JOINED_XLSX, SPECIAL_COPYRIGHT, SPECIAL_INPUT_CSV]


# ── CHART STYLE ──────────────────────────────────────────────────────────────

R = 16; GAP = 5; PAD = 2; STROKE = 0.5; SCALE = 3.0
FONT = "Helvetica Neue, Helvetica, Arial, sans-serif"; FONT_SZ = 8.5
STEP = R * 2 + GAP  # 37


# ── DISTRIBUTION MAP ─────────────────────────────────────────────────────────

ATLAS_ID          = 11
CANVAS_W          = 1200
CANVAS_H          = 800
MAP_SCALE         = 3
LV03_XMIN         = 478000
LV03_XMAX         = 836000
LV03_YMIN         = 68000
LV03_YMAX         = 299000
SMOOTH_SIGMA_EDGE = 12.0
SMOOTH_THRESHOLD  = 0.40
BORDER_THICKNESS  = 2
LAKE_GENEVA_IDX   = 27
LAKE_COLOR        = (195, 195, 195)
FILL_COLOR        = (60, 60, 60)


def wgs84_to_lv03(lon, lat):
    phi = (lat * 3600 - 169028.66) / 10000
    lam = (lon * 3600 - 26782.5) / 10000
    x = (200147.07 + 308807.95 * phi + 3745.25 * lam ** 2 + 76.63 * phi ** 2
         - 194.56 * lam ** 2 * phi + 119.79 * phi ** 3)
    y = (600072.37 + 211455.93 * lam - 10938.51 * lam * phi
         - 0.36 * lam * phi ** 2 - 44.54 * lam ** 3)
    return x, y


def lv03_to_px(e, n, w, h):
    px = int((e - LV03_XMIN) / (LV03_XMAX - LV03_XMIN) * w)
    py = int((1 - (n - LV03_YMIN) / (LV03_YMAX - LV03_YMIN)) * h)
    return px, py


def ring_to_px(ring, w, h):
    return np.array([lv03_to_px(c[0], c[1], w, h) for c in ring], dtype=np.int32)


def smooth_ring(pts, sigma=2.5):
    x = uniform_filter1d(pts[:, 0].astype(float), size=int(sigma * 6), mode='wrap')
    y = uniform_filter1d(pts[:, 1].astype(float), size=int(sigma * 6), mode='wrap')
    return np.stack([x, y], axis=1).astype(np.int32)


def fetch_json(url):
    with urlopen(url, timeout=20) as r:
        return json.loads(r.read())


def fetch_taxon_id(url):
    with urlopen(url, timeout=20) as r:
        html = r.read().decode('utf-8', errors='replace')
    m = re.search(r'staticTaxonId\s*=\s*(\d+)', html)
    if not m:
        raise ValueError(f"staticTaxonId not found in {url}")
    return int(m.group(1))


def generate_map(taxon_id, out_path, borders, lakes):
    W = CANVAS_W * MAP_SCALE
    H = CANVAS_H * MAP_SCALE
    url = (f"https://obs.infoflora.ch/rest/v5/communities/0/public"
           f"/atlases/{ATLAS_ID}/names/{taxon_id}/units.geojson?with-legend=true")
    obs = fetch_json(url)
    n_pts = len(obs['features'])
    print(f"  {n_pts} grid points")
    if n_pts == 0:
        os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
        blank = np.full((CANVAS_H, CANVAS_W, 3), 255, dtype=np.uint8)
        cv2.imwrite(out_path, blank, [cv2.IMWRITE_JPEG_QUALITY, 95])
        return False
    lf_gva = lakes['features'][LAKE_GENEVA_IDX]
    img = np.full((H, W, 3), 255, dtype=np.uint8)
    for lf in lakes['features']:
        geom = lf['geometry']
        rings_px = ([ring_to_px(geom['coordinates'][0], W, H)]
                    if geom['type'] == 'Polygon'
                    else [ring_to_px(p[0], W, H) for p in geom['coordinates']])
        for r in rings_px:
            cv2.fillPoly(img, [r], (255, 255, 255))
            cv2.polylines(img, [r], True, LAKE_COLOR, 1)
    border_coords  = borders['features'][0]['geometry']['coordinates'][0]
    border_mask_hi = np.zeros((H, W), dtype=np.uint8)
    cv2.fillPoly(border_mask_hi, [ring_to_px(border_coords, W, H)], 255)
    cv2.fillPoly(border_mask_hi, [ring_to_px(lf_gva['geometry']['coordinates'][0], W, H)], 255)
    border_cnts, _ = cv2.findContours(border_mask_hi, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    border_cnt     = max(border_cnts, key=cv2.contourArea)
    border_cnt_sm  = smooth_ring(border_cnt[:, 0, :], sigma=2.0).reshape(-1, 1, 2)
    cv2.polylines(img, [border_cnt_sm], True, (0, 0, 0), BORDER_THICKNESS * MAP_SCALE)
    CELL_SIZE = 5000
    half_w = int(CELL_SIZE / (LV03_XMAX - LV03_XMIN) * W) // 2
    half_h = int(CELL_SIZE / (LV03_YMAX - LV03_YMIN) * H) // 2
    grid = {}
    for f in obs['features']:
        lon, lat = f['geometry']['coordinates']
        nx, e    = wgs84_to_lv03(lon, lat)
        gc = int(round((e         - LV03_XMIN) / CELL_SIZE))
        gr = int(round((LV03_YMAX - nx)        / CELL_SIZE))
        grid[(gc, gr)] = lv03_to_px(e, nx, W, H)
    cell_mask = np.zeros((H, W), dtype=np.uint8)
    for (px, py) in grid.values():
        x1 = max(0, px - half_w); x2 = min(W, px + half_w)
        y1 = max(0, py - half_h); y2 = min(H, py + half_h)
        cell_mask[y1:y2, x1:x2] = 255
    cell_mask = cv2.dilate(cell_mask, np.ones((3, 3), np.uint8), iterations=1)
    blurred = gaussian_filter(cell_mask.astype(np.float32) / 255.0, sigma=SMOOTH_SIGMA_EDGE)
    smooth  = (blurred > SMOOTH_THRESHOLD).astype(np.uint8) * 255
    smooth  = cv2.bitwise_and(smooth, border_mask_hi)
    lake_mask = np.zeros((H, W), dtype=np.uint8)
    for lf in lakes['features']:
        geom = lf['geometry']
        if geom['type'] == 'Polygon':
            cv2.fillPoly(lake_mask, [ring_to_px(geom['coordinates'][0], W, H)], 255)
        else:
            for poly in geom['coordinates']:
                cv2.fillPoly(lake_mask, [ring_to_px(poly[0], W, H)], 255)
    smooth[lake_mask > 0] = 0
    img[smooth > 0] = FILL_COLOR
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    out = cv2.resize(img, (CANVAS_W, CANVAS_H), interpolation=cv2.INTER_AREA)
    cv2.imwrite(out_path, out, [cv2.IMWRITE_JPEG_QUALITY, 95])
    print(f"  saved: {out_path}")
    return True


# ── PHOTO DOWNLOAD ────────────────────────────────────────────────────────────

LANG = "en"
DELAY = 1.0

_HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def create_placeholder(path, taxon):
    img = Image.new("RGB", (800, 600), "white")
    draw = ImageDraw.Draw(img)
    try:
        font  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 55)
        small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 28)
    except Exception:
        font  = ImageFont.load_default()
        small = ImageFont.load_default()
    text = "Photo not available"
    bbox = draw.textbbox((0, 0), text, font=font)
    x = (800 - (bbox[2] - bbox[0])) // 2
    y = (600 - (bbox[3] - bbox[1])) // 2 - 40
    draw.text((x, y), text, fill=(180, 0, 0), font=font)
    draw.text((40, 420), taxon[:80], fill=(0, 0, 0), font=small)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    img.save(path, quality=95)


def validate_taxon(name):
    if not name:
        return False
    n = name.strip()
    if re.search(r"\b(taxon|rang|familie|checklist|infoflora)\b", n, re.I):
        return False
    if not re.search(r"[A-Z][^\s]+\s+[^\s]+", n):
        return False
    return True



def name_to_slug(name):
    name = name.strip()
    lower = name.lower()
    if "superaggr" in lower:
        tipo = "superaggr"
    elif "aggr" in lower:
        tipo = "aggr"
    elif re.search(r"\bsubsp\.?\b", name, re.I):
        tipo = "subsp"
    elif re.search(r"\bhybrid\b", name, re.I):
        tipo = "hybrid"
    else:
        tipo = "sp"
    clean = re.sub(r"\s*\(.*?\)", "", name)
    clean = re.sub(r"\s+(?:Mill\.|L\.|DC\.|Pers\.|Hoffm\.|Scop\.)", "", clean)
    clean = re.sub(r"\s+", " ", clean).strip()
    if tipo in ("aggr", "superaggr"):
        p = clean.split()
        if len(p) >= 2:
            g   = re.sub(r"[^a-zA-Z]", "", p[0]).lower()
            s   = re.sub(r"[^a-zA-Z-]", "", p[1]).lower()
            suf = "-aggr" if tipo == "aggr" else "-superaggr"
            return f"{g}-{s}{suf}", tipo
    if tipo == "subsp":
        m = re.search(r"([A-Za-z-]+)\s+([a-z-]+)\s+.*?subsp\.?\s+([a-z-]+)", name, re.I)
        if m:
            return f"{m.group(1).lower()}-{m.group(2).lower()}-subsp-{m.group(3).lower()}", tipo
        p = clean.split()
        for i, x in enumerate(p):
            if x.lower().startswith("subsp") and i + 1 < len(p):
                return f"{p[0].lower()}-{p[1].lower()}-subsp-{p[i+1].lower()}", tipo
    p = clean.split()
    if len(p) < 2:
        return (p[0].lower(), "invalid")
    g = re.sub(r"[^\w-]", "", p[0], flags=re.UNICODE).lower()
    s = re.sub(r"[^\w×-]", "", p[1], flags=re.UNICODE).lower()
    return f"{g}-{s}", tipo


def fetch_page(url):
    try:
        r = requests.get(url, headers=_HTTP_HEADERS, timeout=15)
        return r if r.status_code == 200 else None
    except Exception:
        return None


def extract_original_url(imgproxy_url):
    try:
        path = urlparse(imgproxy_url).path
        b64  = path.split("/")[-1]
        b64  = re.sub(r"\.[a-z]+$", "", b64)
        padding = 4 - len(b64) % 4
        if padding != 4:
            b64 += "=" * padding
        return base64.urlsafe_b64decode(b64).decode("utf-8")
    except Exception:
        return imgproxy_url


def clean_credit(credit):
    m = re.search(r'©\s*(?:Copyright\s*)?(.+)', credit)
    if m:
        return "© " + m.group(1).strip()
    return credit.strip()


# URL patterns identifying botanical illustrations and anatomical sections to exclude
_EXCLUDED_URL_PARTS  = ("/dessins/", "/species/anatomy/")
# Editorial credits indicating illustrations (not photos)
_EXCLUDED_PUBLISHERS = ("Flora Vegetativa",)


def _url_accettabile(url):
    return not any(p in url for p in _EXCLUDED_URL_PARTS)


def _title_acceptable(title):
    return not any(p in title for p in _EXCLUDED_PUBLISHERS)


def _extract_gallery_candidates(soup):
    """Returns a sorted list of (url, title) from the gallery, excluding illustrations/anatomical sections."""
    seen, result = set(), []
    for a in soup.select("#info-gallery a[title]"):
        title = a.get("title", "")
        href  = a.get("href", "")
        url   = a.get("original-image") or (extract_original_url(href) if href else "")
        if (url and "infoflora" in url and url not in seen
                and _url_accettabile(url) and _title_acceptable(title)):
            seen.add(url)
            result.append((url, title))
    return result


def save_and_verify(url, path, taxon, output_dir):
    try:
        img = requests.get(url, headers=_HTTP_HEADERS, timeout=30)
        if img.status_code == 200:
            os.makedirs(os.path.dirname(path) or output_dir, exist_ok=True)
            with open(path, "wb") as f:
                f.write(img.content)
            try:
                pil = Image.open(path)
                pil.verify()
                pil = Image.open(path)   # verify() exhausts the file object, reopening is necessary
                arr = np.array(pil.convert("RGB"))
                if float(((arr > 230).all(axis=2)).mean()) > 0.70:
                    create_placeholder(path, taxon)
                    return "WHITE"
                return True
            except Exception:
                create_placeholder(path, taxon)
                return False
    except Exception:
        pass
    return None


def extract_photo(soup, gallery):
    candidates = []
    for url, title in gallery:
        if "Flora Helvetica" in title:
            priority = 0 if "Konrad Lauber" in title else 1
        else:
            priority = 2
        candidates.append((priority, url, title))
    if candidates:
        candidates.sort(key=lambda x: x[0])
        _, img_url, credit = candidates[0]
        return img_url, credit, len(candidates)
    best = None; best_score = -999
    for i in soup.find_all("img"):
        src = i.get("src", ""); alt = (i.get("alt") or "").lower()
        if not src or not _url_accettabile(src):
            continue
        if i.find_parent(id="anatomy-gallery"):
            continue
        score = 0
        if "imgproxy" in src:  score += 5
        if "jpg" in src:       score += 2
        if "logo" in src or "icon" in src: score -= 10
        if "©" in i.get("alt", ""): score += 4
        if len(alt) > 20:      score += 2
        if score > best_score:
            best_score = score; best = i
    if best:
        return best.get("src", ""), best.get("alt", ""), 0
    return "", "", 0


def download_photo(slug, name, output_dir, candidates_dir):
    url = f"https://www.infoflora.ch/{LANG}/flora/{slug}.html"
    r   = fetch_page(url)
    if not r:
        p = os.path.join(output_dir, f"{slug}.jpg")
        create_placeholder(p, name)
        return p, "NO_PAGE", "PAGE_NOT_FOUND"
    soup           = BeautifulSoup(r.text, "html.parser")
    all_candidates = _extract_gallery_candidates(soup)
    img_url, credit, n_photos = extract_photo(soup, all_candidates)
    credit = clean_credit(credit)
    if not img_url or "infoflora" not in img_url:
        p = os.path.join(output_dir, f"{slug}.jpg")
        create_placeholder(p, name)
        return p, "NO_IMAGE", "IMAGE_NOT_FOUND"

    if n_photos <= 1 or "Flora Helvetica" in credit:
        p      = os.path.join(output_dir, f"{slug}.jpg")
        result = save_and_verify(img_url, p, name, output_dir)
        if result is True:      return p, credit, "OK"
        elif result is False:   return p, credit, "INVALID_IMAGE"
        elif result == "WHITE": return p, credit, "WHITE_BG"
        else:
            create_placeholder(p, name)
            return p, "ERROR", "IMAGE_DOWNLOAD_ERROR"

    if not all_candidates:
        p = os.path.join(output_dir, f"{slug}.jpg")
        create_placeholder(p, name)
        return p, "NO_IMAGE", "IMAGE_NOT_FOUND"

    n = len(all_candidates)
    if n <= 4:
        subset = all_candidates
    else:
        indices = sorted({0, n // 3, 2 * n // 3, n - 1})
        subset  = [all_candidates[i] for i in indices]
    candidate_dir = os.path.join(candidates_dir, slug)
    os.makedirs(candidate_dir, exist_ok=True)
    for idx, (u, t) in enumerate(subset):
        save_and_verify(u, os.path.join(candidate_dir, f"{idx+1:02d}_{slug}.jpg"), name, output_dir)
    main_url, main_credit = all_candidates[-1]
    main_credit = clean_credit(main_credit)
    p      = os.path.join(output_dir, f"{slug}.jpg")
    result = save_and_verify(main_url, p, name, output_dir)
    if result is True:      return p, main_credit, "OK_MULTIPLE"
    elif result is False:   return p, main_credit, "INVALID_IMAGE"
    elif result == "WHITE": return p, main_credit, "WHITE_BG"
    else:
        create_placeholder(p, name)
        return p, "ERROR", "IMAGE_DOWNLOAD_ERROR"


def _dp_leggi_xlsx(path):
    wb  = openpyxl.load_workbook(path)
    ws  = wb.active
    hi  = find_xlsx_header_row(ws)
    all_rows = list(ws.iter_rows(values_only=True))
    hdrs = [str(v).strip() if v else "" for v in all_rows[hi]]
    def ci(name):
        try: return next(i for i, h in enumerate(hdrs) if h.lower() == name.lower())
        except StopIteration: return None
    nc, rc, fc, ic, ac = ci("Taxonname"), ci("Rang"), ci("Familie"), ci("Indigenat"), ci("Aggregat")
    rows = []
    for idx, row in enumerate(all_rows[hi + 1:], start=hi + 1):
        name = str(row[nc] or "").strip() if nc is not None else ""
        if not validate_taxon(name):
            continue
        rows.append({
            "indice":    idx,
            "taxonname": name,
            "rang":      str(row[rc] or "") if rc is not None else "",
            "famiglia":  str(row[fc] or "") if fc is not None else "",
            "indigenat": str(row[ic] or "") if ic is not None else "",
            "aggregato": str(row[ac] or "").strip() if ac is not None else "",
        })
    return rows


def _dp_leggi_csv(path):
    rows = []
    with open(path, encoding="utf-8-sig") as f:
        content = f.read()
    sep    = ";" if content.count(";") > content.count(",") else ","
    reader = csv.DictReader(content.splitlines(), delimiter=sep)
    col    = reader.fieldnames[0]
    for i, row in enumerate(reader):
        name = row[col].strip()
        if not validate_taxon(name):
            continue
        rows.append({"indice": i, "taxonname": name, "rang": "sp", "aggregato": ""})
    return rows


def load_existing_copyright(path_csv):
    credits = {}
    if not os.path.exists(path_csv):
        return credits
    with open(path_csv, encoding="utf-8") as f:
        for row in csv.DictReader(f, delimiter=";"):
            file_val = row.get("File", ""); credit = row.get("Credit", "")
            if file_val and credit and credit not in ("", "RESUMED"):
                slug = os.path.splitext(os.path.basename(file_val))[0]
                credits[slug] = credit
    return credits


def _download_batch(input_path, output_dir, copyright_csv, candidates_dir):
    species = _dp_leggi_xlsx(input_path) if input_path.endswith(".xlsx") else _dp_leggi_csv(input_path)
    cache  = {}
    crediti_esistenti = load_existing_copyright(copyright_csv)
    results = []
    for i, sp in enumerate(species):
        slug     = name_to_slug(sp["taxonname"])[0]
        p_atteso = os.path.join(output_dir, f"{slug}.jpg")
        if os.path.exists(p_atteso) and slug not in cache:
            print(f"{i} ↷ already present: {p_atteso}")
            credit_noto = crediti_esistenti.get(slug, "")
            cache[slug] = (p_atteso, credit_noto, "RESUMED")
        if slug in cache:
            p, credit, status = cache[slug]
            src = cache[slug][0]
            if os.path.abspath(src) != os.path.abspath(p_atteso):
                shutil.copy2(src, p_atteso)
            results.append([sp["indice"], sp["taxonname"], sp["rang"],
                               sp.get("aggregato", ""), p_atteso, credit, status])
            continue
        print(f"{i} {sp['taxonname']}")
        p, credit, status = download_photo(slug, sp["taxonname"], output_dir, candidates_dir)
        cache[slug] = (p, credit, status)
        results.append([sp["indice"], sp["taxonname"], sp["rang"],
                          sp.get("aggregato", ""), p, credit, status])
        time.sleep(DELAY)
    with open(copyright_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Idx", "Name", "Rank", "Aggregato", "File", "Credit", "Status"])
        w.writerows(results)
    print("\nDONE")


# ── DATA SCRAPING ─────────────────────────────────────────────────────────────

_DZ_LABEL_MAP = {
    "feuchtezahl":          "F",
    "reaktionszahl":        "R",
    "nährstoffzahl":        "N",
    "lichtzahl":            "L",
    "temperaturzahl":       "T",
    "kontinentalitätszahl": "K",
    "salzzeichen":          "Salt",
}


def _split_indici(parte):
    return re.findall(r"\d[^\d]*", parte)


def _parse_compact(compact):
    m = re.match(r"^([^-]+)-([^.]+)(\..*)?$", compact.strip())
    if not m:
        return None
    frn_raw = m.group(1)
    ltk_raw = m.group(2)
    rest    = m.group(3) or ""
    frn = _split_indici(frn_raw)
    ltk = _split_indici(ltk_raw)
    if len(frn) < 3 or len(ltk) < 3:
        return None
    lifeform = ""
    if rest:
        # rest starts with ".", e.g. ".k-t.2n=14" or ".h"
        inner    = rest[1:]          # strip the leading dot
        lifeform = inner.split(".")[0]  # first segment is the lifeform code
    return {
        "F": frn[0], "R": frn[1], "N": frn[2],
        "L": ltk[0], "T": ltk[1], "K": ltk[2],
        "Lebensform": lifeform,
    }


def _extract_compact(soup):
    for h4 in soup.find_all("h4"):
        if "Flora Helvetica" in h4.get_text() and "Zeigerwerte" in h4.get_text():
            BLOCK = {"div", "table", "p", "ul", "ol", "h1", "h2", "h3", "h4", "h5", "h6", "section"}
            parts = []
            for sib in h4.next_siblings:
                if getattr(sib, "name", None) in BLOCK:
                    break
                parts.append(sib.get_text() if hasattr(sib, "get_text") else str(sib))
            return "".join(parts).strip()
    return ""


def extract_zeigerwerte(html):
    soup    = BeautifulSoup(html, "html.parser")
    compact = _extract_compact(soup)
    vals    = {}
    table   = soup.find("table", attrs={"aria-describedby": "ecology-landolt-legend"})
    if table:
        for row in table.find_all("tr"):
            tds = row.find_all("td")
            for i in range(0, len(tds) - 1, 2):
                label = tds[i].get_text(strip=True).lower()
                value = tds[i + 1].get_text(strip=True)
                for key, col in _DZ_LABEL_MAP.items():
                    if key in label:
                        vals[col] = value
                        break
    if not all(k in vals for k in ("F", "R", "N", "L", "T", "K")):
        if compact:
            parsed = _parse_compact(compact)
            if parsed:
                return parsed
        return None
    lifeform = ""
    if compact:
        parsed = _parse_compact(compact)
        if parsed:
            lifeform = parsed["Lebensform"]
        elif "." in compact:
            # fallback: Lebensform is always after the last dot
            lifeform = re.sub(r"[^a-zA-Z*]", "", compact.rsplit(".", 1)[-1])
    return {
        "F": vals["F"], "R": vals["R"], "N": vals["N"],
        "L": vals["L"], "T": vals["T"], "K": vals["K"],
        "Lebensform": lifeform,
    }


def extract_flowering(html):
    soup = BeautifulSoup(html, "html.parser")
    for h4 in soup.find_all("h4"):
        if "fioritura" in h4.get_text(strip=True).lower():
            for sib in h4.next_siblings:
                text = (sib.get_text(strip=True) if hasattr(sib, "get_text") else str(sib)).strip()
                if text:
                    m = re.search(r"(\d{1,2})\s*[-–]\s*(\d{1,2})", text)
                    if m:
                        return int(m.group(1)), int(m.group(2))
                    break
    return None


def extract_red_list(html):
    soup = BeautifulSoup(html, "html.parser")
    img  = soup.find("img", src=re.compile(r"/redlist/"))
    if img:
        m = re.search(r"/([A-Z/]+)\.png", img["src"])
        if m:
            return m.group(1).split("/")[-1]
    for h4 in soup.find_all("h4"):
        if "IUCN" in h4.get_text():
            nxt = h4.find_next_sibling("p")
            if nxt:
                m = re.search(r"\(([A-Z/]+)\)", nxt.get_text())
                if m:
                    return m.group(1)
    return "N/D"


# ── GENERATE FLOWERING CHART ──────────────────────────────────────────────────

_MESI    = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_FI_COLS = 3
_FI_ROWS = 4
_FI_GAP  = 5
_FI_W    = _FI_COLS * STEP - _FI_GAP + PAD * 2   # 110
_FI_H    = round(_FI_W * 19 / 28.0)               # 75  (ratio ≈ box 28.2×19 mm)
_FI_CELL_W = STEP
_FI_CELL_H = _FI_H / _FI_ROWS
_FI_FONT_SZ = round(_FI_CELL_H * 0.36)


def _svg_fioritura(start, end):
    els = []
    for i, name in enumerate(_MESI):
        m   = i + 1
        row = (m - 1) // 3
        col = (m - 1) % 3
        cx  = PAD + col * STEP + R
        cy  = row * _FI_CELL_H + _FI_CELL_H / 2 + _FI_FONT_SZ * 0.37
        on  = start <= m <= end
        els.append(
            f'<text x="{cx}" y="{cy:.1f}" font-family="{FONT}" '
            f'font-size="{_FI_FONT_SZ}" font-weight="{"700" if on else "300"}" '
            f'fill="{"#000000" if on else "#BBBBBB"}" '
            f'text-anchor="middle">{name}</text>'
        )
    body = "\n  ".join(els)
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{_FI_W}" height="{_FI_H}" viewBox="0 0 {_FI_W} {_FI_H}">\n  {body}\n</svg>\n'
    )


def _generate_flowering(slug, start, end, out_dir=os.path.join(OUTPUT_DIR, "fioritura")):
    os.makedirs(out_dir, exist_ok=True)
    jpg_path  = os.path.join(out_dir, f"{slug}.jpg")
    png_bytes = cairosvg.svg2png(bytestring=_svg_fioritura(start, end).encode(), scale=SCALE)
    png_img   = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
    bg        = Image.new("RGB", png_img.size, (255, 255, 255))
    bg.paste(png_img, mask=png_img.split()[3])
    bg.save(jpg_path, "JPEG", quality=95)
    return jpg_path


# ── GENERATE IUCN BARS ────────────────────────────────────────────────────────

CATEGORIE = ["EX", "CR", "EN", "VU", "NT", "LC"]
COLORI    = {
    "EX": "#E8380D", "CR": "#E86A0D", "EN": "#E8A80D",
    "VU": "#E8D40D", "NT": "#9DC40D", "LC": "#4AAF0D",
}
_IUCN_W = len(CATEGORIE) * STEP - GAP + PAD * 2
_IUCN_H = R * 2 + PAD * 2


def _svg_iucn(status):
    els = []
    for i, cat in enumerate(CATEGORIE):
        cx = PAD + i * STEP + R; cy = PAD + R; on = cat == status.upper()
        if on:
            els.append(f'<circle cx="{cx}" cy="{cy}" r="{R}" fill="{COLORI[cat]}"/>')
            txt_fill = "white"; txt_weight = "600"
        else:
            els.append(
                f'<circle cx="{cx}" cy="{cy}" r="{R}" fill="none" '
                f'stroke="#000000" stroke-width="{STROKE}"/>'
            )
            txt_fill = "#000000"; txt_weight = "400"
        els.append(
            f'<text x="{cx}" y="{cy + FONT_SZ * 0.37:.1f}" font-family="{FONT}" '
            f'font-size="{FONT_SZ}" font-weight="{txt_weight}" fill="{txt_fill}" '
            f'text-anchor="middle">{cat}</text>'
        )
    body = "\n  ".join(els)
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{_IUCN_W}" height="{_IUCN_H}" viewBox="0 0 {_IUCN_W} {_IUCN_H}">\n  {body}\n</svg>\n'
    )


def _generate_iucn(status, out_dir=os.path.join(OUTPUT_DIR, "iucn_bars")):
    os.makedirs(out_dir, exist_ok=True)
    label    = status.upper() if status.upper() in CATEGORIE else "ND"
    jpg_path = os.path.join(out_dir, f"{label}.jpg")
    png_bytes = cairosvg.svg2png(bytestring=_svg_iucn(label).encode(), scale=SCALE)
    png_img   = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
    bg        = Image.new("RGB", png_img.size, (255, 255, 255))
    bg.paste(png_img, mask=png_img.split()[3])
    bg.save(jpg_path, "JPEG", quality=95)
    return jpg_path


# ── GENERATE ZEIGERWERTE IMAGES ───────────────────────────────────────────────

_ZW_INDICI  = [["F", "R", "N"], ["L", "T", "K"]]
_ZW_W, _ZW_H = 286, 102
_ZW_GAP  = 3; _ZW_PAD  = 0.5
_ZW_COLS = 3; _ZW_ROWS = 2
_ZW_CELL_W  = (_ZW_W - 2 * _ZW_PAD - (_ZW_COLS - 1) * _ZW_GAP) / _ZW_COLS
_ZW_CELL_H  = (_ZW_H - 2 * _ZW_PAD - (_ZW_ROWS - 1) * _ZW_GAP) / _ZW_ROWS
_ZW_LABEL_H = _ZW_CELL_H * 0.36
_ZW_STROKE  = 0.5
_ZW_BLACK   = "#000000"
_ZW_GRAY    = "#666666"
_ZW_LABEL_SZ = round(_ZW_LABEL_H * 0.58)
_ZW_VALUE_SZ = round((_ZW_CELL_H - _ZW_LABEL_H) * 0.60)


def _is_extreme(val):
    for ch in val:
        if ch.isdigit():
            return ch in ("1", "5")
    return False


def _svg_zeigerwerte(vals):
    els = []
    for r_idx, row in enumerate(_ZW_INDICI):
        for c_idx, key in enumerate(row):
            x  = _ZW_PAD + c_idx * (_ZW_CELL_W + _ZW_GAP)
            y  = _ZW_PAD + r_idx * (_ZW_CELL_H + _ZW_GAP)
            cx = x + _ZW_CELL_W / 2
            val = str(vals.get(key, "—"))
            els.append(
                f'<rect x="{x:.1f}" y="{y:.1f}" width="{_ZW_CELL_W:.1f}" '
                f'height="{_ZW_LABEL_H:.1f}" fill="#F0F0F0"/>'
            )
            els.append(
                f'<rect x="{x:.1f}" y="{y:.1f}" width="{_ZW_CELL_W:.1f}" '
                f'height="{_ZW_CELL_H:.1f}" fill="none" stroke="{_ZW_BLACK}" '
                f'stroke-width="{_ZW_STROKE}"/>'
            )
            els.append(
                f'<line x1="{x:.1f}" y1="{y + _ZW_LABEL_H:.1f}" '
                f'x2="{x + _ZW_CELL_W:.1f}" y2="{y + _ZW_LABEL_H:.1f}" '
                f'stroke="{_ZW_BLACK}" stroke-width="{_ZW_STROKE * 0.6:.2f}"/>'
            )
            els.append(
                f'<text x="{cx:.1f}" y="{y + _ZW_LABEL_H * 0.74:.1f}" '
                f'font-family="{FONT}" font-size="{_ZW_LABEL_SZ}" font-weight="400" '
                f'fill="{_ZW_GRAY}" text-anchor="middle" letter-spacing="1">{key}</text>'
            )
            extreme    = _is_extreme(val)
            val_font   = FONT if extreme else "Helvetica Neue UltraLight, Helvetica Neue, Arial, sans-serif"
            val_weight = "700" if extreme else "200"
            val_y      = y + _ZW_LABEL_H + (_ZW_CELL_H - _ZW_LABEL_H) * 0.68
            els.append(
                f'<text x="{cx:.1f}" y="{val_y:.1f}" font-family="{val_font}" '
                f'font-size="{_ZW_VALUE_SZ}" font-weight="{val_weight}" '
                f'fill="{_ZW_BLACK}" text-anchor="middle">{val}</text>'
            )
    body = "\n  ".join(els)
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{_ZW_W}" height="{_ZW_H}" viewBox="0 0 {_ZW_W} {_ZW_H}">\n  {body}\n</svg>\n'
    )


def _generate_zeigerwerte(slug, vals, out_dir=os.path.join(OUTPUT_DIR, "zeigerwerte")):
    os.makedirs(out_dir, exist_ok=True)
    jpg_path  = os.path.join(out_dir, f"{slug}.jpg")
    png_bytes = cairosvg.svg2png(bytestring=_svg_zeigerwerte(vals).encode(), scale=SCALE)
    png_img   = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
    bg        = Image.new("RGB", png_img.size, (255, 255, 255))
    bg.paste(png_img, mask=png_img.split()[3])
    bg.save(jpg_path, "JPEG", quality=95)
    return jpg_path


# ── LOGGING ───────────────────────────────────────────────────────────────────

class _Tee:
    def __init__(self, path):
        self._file   = open(path, "a", encoding="utf-8")
        self._stdout = sys.stdout

    def write(self, data):
        self._stdout.write(data)
        self._file.write(data)

    def flush(self):
        self._stdout.flush()
        self._file.flush()

    def close(self):
        self._file.close()


def _start_log():
    tee = _Tee(LOG_PATH)
    sys.stdout = tee
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'#' * 60}")
    print(f"#  Run started: {ts}")
    print(f"{'#' * 60}")
    return tee


def _stop_log(tee):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n#  Run ended: {ts}")
    sys.stdout = tee._stdout
    tee.close()


# ── PIPELINE STEPS ────────────────────────────────────────────────────────────

def banner(step, title):
    print()
    print("=" * 60)
    print(f"  STEP {step} — {title}")
    print("=" * 60)


def read_copyright_csv(path):
    data = {}
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f, delimiter=";"):
            name = row.get("Name", "").strip()
            if name:
                data[name] = dict(row)
    return data


def find_xlsx_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).lower().strip() for v in row if v is not None]
        if any(v in HEADER_KEYWORDS for v in vals):
            return i
    return 0


def _read_xlsx_species(path):
    wb         = openpyxl.load_workbook(path)
    ws         = wb.active
    header_idx = find_xlsx_header_row(ws)
    rows       = list(ws.iter_rows(values_only=True))
    headers    = [str(v).strip() if v else "" for v in rows[header_idx]]
    try:
        name_col = next(
            i for i, h in enumerate(headers)
            if "taxonname" in h.lower() or "nom du taxon" in h.lower()
        )
    except StopIteration:
        raise ValueError("Column 'Taxonname' not found in xlsx.")

    # SISF/Info Flora exports have a units/legend row right after the header,
    # with no Taxonname value. Only skip it if it's actually empty, so a
    # plain header+data xlsx (no legend row) doesn't lose its first species.
    data_start = header_idx + 1
    if data_start < len(rows):
        next_row  = rows[data_start]
        next_name = str(next_row[name_col]).strip() if next_row[name_col] else ""
        if not next_name or next_name == "None":
            data_start += 1

    data = {}
    for row in rows[data_start:]:
        if all(v is None for v in row):
            continue
        name = str(row[name_col]).strip() if row[name_col] else ""
        if name and name != "None":
            data[name] = [v if v is not None else "" for v in row]
    return headers, data


def step_main_download(input_xlsx):
    banner(1, f"Main download  →  {MAIN_OUTPUT_DIR}/")
    _download_batch(input_xlsx, MAIN_OUTPUT_DIR, MAIN_COPYRIGHT, MAIN_CANDIDATES)


_PLACEHOLDER_STATUSES = {"IMAGE_NOT_FOUND", "PAGE_NOT_FOUND", "NO_PAGE",
                         "NO_IMAGE", "IMAGE_DOWNLOAD_ERROR", "INVALID_IMAGE", "WHITE_BG"}


def _load_existing_data(existing_xlsx):
    """Reads credits and original status values from a previous run's output."""
    data = {}
    if not existing_xlsx or not os.path.exists(existing_xlsx):
        return data
    try:
        wb = openpyxl.load_workbook(existing_xlsx)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if "Taxonname" not in headers:
            return data
        name_idx   = headers.index("Taxonname")
        credit_idx = headers.index("Credit") if "Credit" in headers else None
        status_idx = headers.index("Status") if "Status" in headers else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_idx] or "").strip()
            if not name:
                continue
            data[name] = {
                "credit": str(row[credit_idx] or "").strip() if credit_idx is not None else "",
                "status": str(row[status_idx] or "").strip() if status_idx is not None else "",
            }
        print(f"  Existing data loaded from {existing_xlsx}: {len(data)} species")
    except Exception as e:
        print(f"  Warning: unable to read data from {existing_xlsx}: {e}")
    return data


def step_join(xlsx_headers, xlsx_data, existing_xlsx=None):
    banner(2, "Joining tables")
    csv_data      = read_copyright_csv(MAIN_COPYRIGHT)
    existing_data = _load_existing_data(existing_xlsx)
    wb_out   = openpyxl.Workbook()
    ws_out   = wb_out.active
    ws_out.title = "Joined"
    ws_out.append(xlsx_headers + CSV_EXTRA_COLS)
    for cell in ws_out[1]:
        cell.font = Font(bold=True)
    matched   = 0
    unmatched = []
    for name, xlsx_row in xlsx_data.items():
        csv_row = csv_data.get(name)
        if csv_row:
            credit = csv_row.get("Credit", "")
            status = csv_row.get("Status", "")
            prev   = existing_data.get(name, {})
            if not credit:
                credit = prev.get("credit", "")
            if status == "RESUMED" and prev.get("status") in _PLACEHOLDER_STATUSES:
                status = prev["status"]
            extra = [
                csv_row.get("File", ""),
                credit,
                status,
                csv_row.get("Map",  ""),
            ]
            matched += 1
        else:
            extra = [""] * len(CSV_EXTRA_COLS)
            unmatched.append(name)
        ws_out.append(list(xlsx_row) + extra)
    wb_out.save(JOINED_XLSX)
    print(f"Matched   : {matched}")
    print(f"Unmatched : {len(unmatched)}")
    if unmatched:
        print("\nMissed matches:")
        for name in unmatched:
            print(f"  - {name}")
    return unmatched


def step_special_download(missed):
    banner(3, f"Special cases  ({len(missed)} species)  →  {SPECIAL_OUTPUT_DIR}/")
    if not missed:
        print("No missed matches — skipping.")
        return
    with open(SPECIAL_INPUT_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Taxonname"])
        for name in missed:
            w.writerow([name])
    _download_batch(SPECIAL_INPUT_CSV, SPECIAL_OUTPUT_DIR, SPECIAL_COPYRIGHT, SPECIAL_CANDIDATES)


def step_maps(xlsx_data):
    banner(4, f"Generating maps  →  {MAPS_DIR}/")
    print("Downloading CH border and lakes (once)…")
    borders = fetch_json("https://atlas.infoflora.ch/resources/json/borders.geojson")
    lakes   = fetch_json("https://atlas.infoflora.ch/resources/json/lakes.geojson")
    os.makedirs(MAPS_DIR, exist_ok=True)
    maps = {}
    for name in xlsx_data:
        slug     = name_to_slug(name)[0]
        out_path = os.path.join(MAPS_DIR, f"{slug}.jpg")
        if os.path.exists(out_path):
            print(f"  ↷ already present: {out_path}")
            maps[name] = out_path
            continue
        url = f"https://www.infoflora.ch/en/flora/{slug}.html"
        print(f"  {name}")
        try:
            taxon_id = fetch_taxon_id(url)
            generate_map(taxon_id, out_path, borders, lakes)
        except Exception as e:
            print(f"    error: {e} — blank map")
            blank = np.full((CANVAS_H, CANVAS_W, 3), 255, dtype=np.uint8)
            cv2.imwrite(out_path, blank, [cv2.IMWRITE_JPEG_QUALITY, 95])
        maps[name] = out_path
    print(f"Maps generated: {len(maps)} / {len(xlsx_data)}")
    return maps


def step_final_merge(output_xlsx, maps_data=None):
    banner(5, f"Producing final output  →  {output_xlsx}")
    if maps_data is None:
        maps_data = {}
    wb      = openpyxl.load_workbook(JOINED_XLSX)
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]
    name_col_idx   = next(i for i, h in enumerate(headers) if h and "taxonname" in str(h).lower())
    file_col_idx   = next(i for i, h in enumerate(headers) if h == "File")
    credit_col_idx = next(i for i, h in enumerate(headers) if h == "Credit")
    status_col_idx = next(i for i, h in enumerate(headers) if h == "Status")
    map_col_idx    = next((i for i, h in enumerate(headers) if h == "Map"), None)
    special_data   = {}
    if os.path.exists(SPECIAL_COPYRIGHT):
        special_data = read_copyright_csv(SPECIAL_COPYRIGHT)
    newly_matched = []
    still_missing = []
    for row in ws.iter_rows(min_row=2):
        name     = str(row[name_col_idx].value or "").strip()
        file_val = str(row[file_col_idx].value or "").strip()
        if not file_val:
            sp = special_data.get(name)
            if sp:
                row[file_col_idx].value   = sp.get("File",   "")
                row[credit_col_idx].value = sp.get("Credit", "")
                row[status_col_idx].value = sp.get("Status", "")
                newly_matched.append((name, sp.get("Status", "")))
            else:
                still_missing.append(name)
        if map_col_idx is not None and name in maps_data:
            row[map_col_idx].value = maps_data[name]
    wb.save(output_xlsx)
    print(f"Output        : {output_xlsx}")
    print(f"Newly filled  : {len(newly_matched)}")
    print(f"Still missing : {len(still_missing)}")
    if newly_matched:
        print("\nNewly matched:")
        for name, status in newly_matched:
            print(f"  + {name}  [{status}]")
    if still_missing:
        print("\nStill missing:")
        for name in still_missing:
            print(f"  - {name}")


def step_scrape_dati(output_xlsx):
    banner(6, "Fetching ecological data (Zeigerwerte · IUCN · Fioritura)")
    _H = {"User-Agent": "Mozilla/5.0"}
    ZW_COLS  = ["F", "R", "N", "L", "T", "K", "Lebensform"]
    OUT_FIOR = os.path.join(OUTPUT_DIR, "fioritura")

    wb      = openpyxl.load_workbook(output_xlsx)
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]

    for col in ZW_COLS + ["Lista Rossa", "Fioritura", "Fioritura_Start", "Fioritura_End"]:
        if col not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=col).font = Font(bold=True)
            headers.append(col)

    file_col_idx = next(i for i, h in enumerate(headers) if h == "File")

    for row in ws.iter_rows(min_row=2):
        name     = str(row[0].value or "").strip()
        file_val = str(row[file_col_idx].value or "").strip()
        if not file_val:
            row[headers.index("Lista Rossa")].value = "N/D"
            continue

        slug = os.path.splitext(os.path.basename(file_val))[0]
        print(f"  {name}")

        # GET 1: German page → Zeigerwerte
        try:
            r_de  = requests.get(f"https://www.infoflora.ch/de/flora/{slug}.html",
                                 headers=_H, timeout=15)
            dati_zw = extract_zeigerwerte(r_de.text) if r_de.status_code == 200 else None
        except Exception:
            dati_zw = None

        for col in ZW_COLS:
            row[headers.index(col)].value = dati_zw.get(col, "") if dati_zw else ""

        if dati_zw:
            vals = " | ".join(f"{k}={dati_zw[k]}" for k in ["F", "R", "N", "L", "T", "K"])
            print(f"    ZW  → {vals}")
        else:
            print(f"    ZW  → not found")

        time.sleep(0.5)

        # GET 2: Italian page → Lista Rossa + Fioritura
        try:
            r_it = requests.get(f"https://www.infoflora.ch/it/flora/{slug}.html",
                                headers=_H, timeout=15)
            if r_it.status_code == 200:
                stato        = extract_red_list(r_it.text)
                result_fior  = extract_flowering(r_it.text)
            else:
                stato = "N/D"; result_fior = None
        except Exception:
            stato = "N/D"; result_fior = None

        row[headers.index("Lista Rossa")].value = stato
        print(f"    IUCN → {stato}")

        if result_fior:
            start, end = result_fior
            jpg = _generate_flowering(slug, start, end, OUT_FIOR)
            row[headers.index("Fioritura")].value = jpg
            row[headers.index("Fioritura_Start")].value = start
            row[headers.index("Fioritura_End")].value = end
            print(f"    Fior → {start}-{end}  {jpg}")
        else:
            row[headers.index("Fioritura")].value = ""
            row[headers.index("Fioritura_Start")].value = ""
            row[headers.index("Fioritura_End")].value = ""
            print(f"    Fior → not found")

        time.sleep(0.5)

    wb.save(output_xlsx)
    print(f"Ecological data updated in {output_xlsx}")


def step_genera_zeigerwerte_images(output_xlsx):
    banner(7, "Generating Zeigerwerte images")
    wb      = openpyxl.load_workbook(output_xlsx)
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]
    zw_keys      = ["F", "R", "N", "L", "T", "K"]
    file_col_idx = next(i for i, h in enumerate(headers) if h == "File")
    for row in ws.iter_rows(min_row=2, values_only=True):
        file_val = str(row[file_col_idx] or "").strip()
        if not file_val:
            continue
        slug = os.path.splitext(os.path.basename(file_val))[0]
        out  = os.path.join("zeigerwerte", f"{slug}.jpg")
        if os.path.exists(out):
            print(f"  ↷ already present: {out}")
            continue
        vals = {k: str(row[headers.index(k)] or "") for k in zw_keys if k in headers}
        if not any(vals.values()):
            print(f"  — no data for {slug}")
            continue
        _generate_zeigerwerte(slug, vals)
        print(f"  ✓ {out}")


def step_genera_iucn_bars():
    banner(8, "Generating IUCN bar images")
    for status in CATEGORIE:
        out = os.path.join("iucn_bars", f"{status}.jpg")
        if os.path.exists(out):
            print(f"  ↷ already present: {out}")
            continue
        _generate_iucn(status)
        print(f"  ✓ {out}")


def step_cleanup():
    banner(9, "Cleaning up intermediate files")
    for path in INTERMEDIATES:
        if os.path.exists(path):
            os.remove(path)
            print(f"  removed: {path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: venv/bin/python3 phase1.py <input.xlsx>")
        sys.exit(1)

    input_xlsx  = sys.argv[1]
    base        = os.path.splitext(os.path.basename(input_xlsx))[0]
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_xlsx = os.path.join(OUTPUT_DIR, f"{base}_output.xlsx")

    tee = _start_log()
    try:
        print("=" * 60)
        print(f"  PIPELINE — infoflora photo downloader")
        print(f"  Input  : {input_xlsx}")
        print(f"  Output : {output_xlsx}")
        print("=" * 60)

        step_main_download(input_xlsx)

        xlsx_headers, xlsx_data = _read_xlsx_species(input_xlsx)
        missed = step_join(xlsx_headers, xlsx_data, existing_xlsx=output_xlsx)

        step_special_download(missed)

        maps_data = step_maps(xlsx_data)

        step_final_merge(output_xlsx, maps_data)

        step_scrape_dati(output_xlsx)

        step_genera_zeigerwerte_images(output_xlsx)

        step_genera_iucn_bars()

        step_cleanup()

        print()
        print("=" * 60)
        print(f"  PIPELINE COMPLETE  —  {output_xlsx}")
        print("=" * 60)

    except Exception as e:
        print(f"\nPIPELINE ERROR: {e}")
        raise
    finally:
        _stop_log(tee)


if __name__ == "__main__":
    main()
