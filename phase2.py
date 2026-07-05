#!/usr/bin/env python3
"""
phase2.py — Phase 2: production of print-ready labels.

Prerequisites:
  - species_output.xlsx  (output of pipeline.py)
  - Assets in: maps/, zeigerwerte/, fioritura/, iucn_bars/
  - Candidate photos in images_ext_candidates/{slug}/ or images_ext/{slug}.jpg
    (unwanted candidates should be removed by the user before running this script)

Produces:
  - foto/                    — one photo per species renamed by slug
  - <base>_retro.pdf         — A4 portrait, 2×4 labels rotated 90°
  - <base>_fronte.pdf        — A4 portrait, 2×4 photos (mirrored columns for duplex)

Duplex printing (long-edge flip):
  1. Print retro.pdf face-down
  2. Reload (flip along the long edge)
  3. Print fronte.pdf
  → front and back align perfectly

Usage:
  python3 phase2.py species_output.xlsx [--solo-foto] [--lang it|en|de|fr]

  --lang picks the card's text language (Indigenat, Lebensform, flowering
  months); default is "it". Flowering months are only re-rendered in the
  chosen language when phase1 captured the raw start/end months (columns
  Fioritura_Start/Fioritura_End) — older xlsx files without them keep the
  pre-rendered (English) chart.
"""

import sys
import os
import re
import io
import base64
import shutil
import tempfile
import time
from urllib.parse import urlparse

import numpy as np
import requests
import openpyxl
import cairosvg
from bs4 import BeautifulSoup
from PIL import Image, ImageOps
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# ── FONT ──────────────────────────────────────────────────────────────────────
_TTC      = "/System/Library/Fonts/HelveticaNeue.ttc"
_FONT_DIR = os.path.join(BASE_DIR, "fonts")
_INTER    = {
    "HN":        "Inter-Regular.ttf",
    "HN-Bold":   "Inter-Bold.ttf",
    "HN-Light":  "Inter-Light.ttf",
    "HN-LightIt":"Inter-LightItalic.ttf",
}
_INTER_ZIP = "https://github.com/rsms/inter/releases/download/v4.0/Inter-4.0.zip"
_INTER_ZIP_PATHS = {
    "extras/ttf/Inter-Regular.ttf":     "Inter-Regular.ttf",
    "extras/ttf/Inter-Bold.ttf":        "Inter-Bold.ttf",
    "extras/ttf/Inter-Light.ttf":       "Inter-Light.ttf",
    "extras/ttf/Inter-LightItalic.ttf": "Inter-LightItalic.ttf",
}

def _download_inter():
    import urllib.request, zipfile, io as _io
    print(f"Downloading Inter fonts from {_INTER_ZIP} ...")
    os.makedirs(_FONT_DIR, exist_ok=True)
    data = urllib.request.urlopen(_INTER_ZIP).read()
    with zipfile.ZipFile(_io.BytesIO(data)) as z:
        for src, dst in _INTER_ZIP_PATHS.items():
            with open(os.path.join(_FONT_DIR, dst), "wb") as f:
                f.write(z.read(src))
            print(f"  → fonts/{dst}")

def _register_fonts():
    if os.path.exists(_TTC):
        # macOS: use system Helvetica Neue
        for name, idx in [("HN",0),("HN-Bold",1),("HN-Light",7),("HN-LightIt",8)]:
            pdfmetrics.registerFont(TTFont(name, _TTC, subfontIndex=idx))
    else:
        # Other OS: use bundled Inter fonts (auto-download if missing)
        missing = [f for f in _INTER.values()
                   if not os.path.exists(os.path.join(_FONT_DIR, f))]
        if missing:
            _download_inter()
        for name, fname in _INTER.items():
            pdfmetrics.registerFont(TTFont(name, os.path.join(_FONT_DIR, fname)))

_register_fonts()

# ── LABEL CONSTANTS ───────────────────────────────────────────────────────────
CW = 68.2 * mm   # label width
CH = 99.0 * mm   # label height

COL_NAME   = "Taxonname"
COL_FAMILY = "Familie"
COL_ORIGIN  = "Indigenat CH"
COL_LIFEFORM = "Lebensform"
COL_MAP    = "Map"
COL_IUCN   = "Lista Rossa"
COL_FLOW   = "Fioritura"
COL_FLOW_START = "Fioritura_Start"
COL_FLOW_END   = "Fioritura_End"
COL_CREDIT = "Credit"

IUCN_DIR = os.path.join(OUTPUT_DIR, "iucn_bars")
ZW_DIR   = os.path.join(OUTPUT_DIR, "zeigerwerte")

# ── FLOWERING CHART (re-rendered per --lang) ──────────────────────────────────
# Same layout/style as phase1._svg_fioritura, parametrized by month labels so
# the chart can be regenerated in the language picked with --lang.
_FIOR_R = 16; _FIOR_GAP = 5; _FIOR_PAD = 2; _FIOR_SCALE = 3.0
_FIOR_FONT = "Helvetica Neue, Helvetica, Arial, sans-serif"
_FIOR_STEP = _FIOR_R * 2 + _FIOR_GAP
_FIOR_COLS, _FIOR_ROWS = 3, 4
_FIOR_W = _FIOR_COLS * _FIOR_STEP - _FIOR_GAP + _FIOR_PAD * 2
_FIOR_H = round(_FIOR_W * 19 / 28.0)
_FIOR_CELL_H = _FIOR_H / _FIOR_ROWS
_FIOR_FONT_SZ = round(_FIOR_CELL_H * 0.36)

_FIOR_CACHE_DIR = os.path.join(OUTPUT_DIR, "fioritura", "_lang")


def _svg_fioritura_lang(start: int, end: int, months: list) -> str:
    els = []
    for i, name in enumerate(months):
        m   = i + 1
        row = (m - 1) // _FIOR_COLS
        col = (m - 1) % _FIOR_COLS
        cx  = _FIOR_PAD + col * _FIOR_STEP + _FIOR_R
        cy  = row * _FIOR_CELL_H + _FIOR_CELL_H / 2 + _FIOR_FONT_SZ * 0.37
        on  = start <= m <= end
        els.append(
            f'<text x="{cx}" y="{cy:.1f}" font-family="{_FIOR_FONT}" '
            f'font-size="{_FIOR_FONT_SZ}" font-weight="{"700" if on else "300"}" '
            f'fill="{"#000000" if on else "#BBBBBB"}" '
            f'text-anchor="middle">{name}</text>'
        )
    body = "\n  ".join(els)
    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{_FIOR_W}" height="{_FIOR_H}" viewBox="0 0 {_FIOR_W} {_FIOR_H}">\n  {body}\n</svg>\n'
    )


def _generate_fioritura_lang(slug: str, start: int, end: int, months: list, out_dir: str) -> str:
    os.makedirs(out_dir, exist_ok=True)
    jpg_path  = os.path.join(out_dir, f"{slug}.jpg")
    png_bytes = cairosvg.svg2png(bytestring=_svg_fioritura_lang(start, end, months).encode(), scale=_FIOR_SCALE)
    png_img   = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
    bg        = Image.new("RGB", png_img.size, (255, 255, 255))
    bg.paste(png_img, mask=png_img.split()[3])
    bg.save(jpg_path, "JPEG", quality=95)
    return jpg_path


def _flowering_image_path(label: dict) -> str | None:
    """Returns the flowering chart path, regenerated in the selected --lang
    when the raw start/end months were captured by phase1. Falls back to the
    pre-rendered image (baked in English) if that data isn't available.
    """
    try:
        start = int(label.get(COL_FLOW_START))
        end   = int(label.get(COL_FLOW_END))
    except (TypeError, ValueError):
        return label.get(COL_FLOW) or None

    slug = slug_from_file(label.get("File"))
    if not slug:
        return label.get(COL_FLOW) or None

    cache_dir  = os.path.join(_FIOR_CACHE_DIR, LANG)
    cache_path = os.path.join(cache_dir, f"{slug}.jpg")
    if not os.path.exists(cache_path):
        _generate_fioritura_lang(slug, start, end, MONTHS_BY_LANG[LANG], cache_dir)
    return cache_path


# ── DRAWN IUCN BAR ────────────────────────────────────────────────────────────
_IUCN_CATS = ["EX", "CR", "EN", "VU", "NT", "LC"]
_IUCN_FILL = {          # colours sampled from the original images
    "EX": (0.910, 0.224, 0.055),
    "CR": (0.914, 0.416, 0.051),
    "EN": (0.910, 0.655, 0.051),
    "VU": (0.906, 0.831, 0.047),
    "NT": (0.612, 0.769, 0.047),
    "LC": (0.290, 0.682, 0.051),
}
_IUCN_GAP_MM = 1.5   # gap between IUCN circles and between zeigerwerte cells


def _draw_iucn_bar(c: canvas.Canvas, active: str,
                   x_left: float, y_bottom: float,
                   total_w: float, bar_h: float):
    """Draws the IUCN bar with circles proportional to the zeigerwerte."""
    n    = len(_IUCN_CATS)
    gap  = _IUCN_GAP_MM * mm
    diam = (total_w - (n - 1) * gap) / n
    r    = diam / 2
    cy   = y_bottom + bar_h / 2

    # Zeigerwerte cell border: 1px in 149px row → ~0.17pt at rendering scale
    lw_inactive = 0.17
    lw_active   = 0.17
    font_pt     = diam * 0.33 / mm * 72 / 25.4   # ~33% of diameter in pt
    font_pt     = max(5.0, min(font_pt, 8.0))

    for i, cat in enumerate(_IUCN_CATS):
        cx = x_left + i * (diam + gap) + r
        is_active = (cat == active.upper())

        if is_active:
            fr, fg, fb = _IUCN_FILL[cat]
            c.setFillColorRGB(fr, fg, fb)
            c.setStrokeColorRGB(fr * 0.75, fg * 0.75, fb * 0.75)
            c.setLineWidth(lw_active)
            c.circle(cx, cy, r, fill=1, stroke=1)
            c.setFillColorRGB(1, 1, 1)
        else:
            c.setFillColorRGB(1, 1, 1)
            c.setStrokeColorRGB(0.12, 0.12, 0.12)
            c.setLineWidth(lw_inactive)
            c.circle(cx, cy, r, fill=1, stroke=1)
            c.setFillColorRGB(0.15, 0.15, 0.15)

        c.setFont("HN", font_pt)
        c.drawCentredString(cx, cy - font_pt * (25.4 / 72) * 0.38 * mm, cat)


# ── ZEIGERWERTE SPLIT INTO 3×2 CELLS ─────────────────────────────────────────
# Original image gaps at 858×306px: col x=281..288, x=569..576 | row y=149..156
_ZW_COL_CUTS = [(0, 281), (289, 569), (577, 858)]
_ZW_ROW_CUTS = [(0, 149), (157, 306)]   # top row (F,R,N) and bottom row (L,T,K)


def _draw_zeigerwerte(c: canvas.Canvas, zw_path: str,
                      x_left: float, y_bottom: float,
                      total_w: float, h: float):
    """Redraws the zeigerwerte as 6 separate cells with gap = _IUCN_GAP_MM."""
    from reportlab.lib.utils import ImageReader
    img = Image.open(zw_path).convert("RGB")

    gap   = _IUCN_GAP_MM * mm
    col_w = (total_w - 2 * gap) / 3
    row_h = (h - gap) / 2

    for ri, (y0, y1) in enumerate(_ZW_ROW_CUTS):
        # ri=0 → top row (goes at the top), ri=1 → bottom row
        y_draw = y_bottom + (1 - ri) * (row_h + gap)
        for ci, (x0, x1) in enumerate(_ZW_COL_CUTS):
            cell = img.crop((x0, y0, x1, y1))
            x_draw = x_left + ci * (col_w + gap)
            c.drawImage(
                ImageReader(cell),
                x_draw, y_draw,
                width=col_w, height=row_h,
                preserveAspectRatio=False,
            )


_PLACEHOLDER_STATUSES = {"IMAGE_NOT_FOUND", "PAGE_NOT_FOUND", "NO_PAGE",
                         "NO_IMAGE", "IMAGE_DOWNLOAD_ERROR", "INVALID_IMAGE", "WHITE_BG"}

# ── TRANSLATIONS ──────────────────────────────────────────────────────────────
# Output language for card text (Indigenat, Lebensform, flowering months).
# Overridden by --lang on the command line; see main().
LANG = "it"
LANG_CHOICES = ("it", "en", "de", "fr")

ORIGIN_IT = {
    "I":   "Indigena",
    "A":   "Archeofita",
    "AC":  "Archeofita coltivata",
    "ni":  "Neo-indigena",
    "N":   "Neofita",
    "NC":  "Neofita coltivata",
    "A/N": "Archeofita / Neofita",
    "I/A": "Indigena / Archeofita",
    "I/N": "Indigena / Neofita",
    "lim": "Solo zone limitrofe",
}

ORIGIN_EN = {
    "I":   "Native",
    "A":   "Archaeophyte",
    "AC":  "Archaeophyte (cultivated)",
    "ni":  "Neo-native",
    "N":   "Neophyte",
    "NC":  "Neophyte (cultivated)",
    "A/N": "Archaeophyte / Neophyte",
    "I/A": "Native / Archaeophyte",
    "I/N": "Native / Neophyte",
    "lim": "Border zones only",
}

ORIGIN_DE = {
    "I":   "Einheimisch",
    "A":   "Archäophyt",
    "AC":  "Archäophyt (kultiviert)",
    "ni":  "Neo-einheimisch",
    "N":   "Neophyt",
    "NC":  "Neophyt (kultiviert)",
    "A/N": "Archäophyt / Neophyt",
    "I/A": "Einheimisch / Archäophyt",
    "I/N": "Einheimisch / Neophyt",
    "lim": "Nur Grenzgebiete",
}

ORIGIN_FR = {
    "I":   "Indigène",
    "A":   "Archéophyte",
    "AC":  "Archéophyte (cultivé)",
    "ni":  "Néo-indigène",
    "N":   "Néophyte",
    "NC":  "Néophyte (cultivé)",
    "A/N": "Archéophyte / Néophyte",
    "I/A": "Indigène / Archéophyte",
    "I/N": "Indigène / Néophyte",
    "lim": "Zones limitrophes uniquement",
}

ORIGIN_BY_LANG = {"it": ORIGIN_IT, "en": ORIGIN_EN, "de": ORIGIN_DE, "fr": ORIGIN_FR}

LIFEFORM_IT = {
    "p":  "fanerofita decidua",
    "i":  "fanerofita sempreverde",
    "n":  "nanofanerofita decidua",
    "j":  "nanofanerofita sempreverde",
    "z":  "camefita lignosa",
    "c":  "camefita erbacea",
    "e":  "epifita",
    "h":  "emicriptofita",
    "g":  "geofita",
    "t":  "terofita",
    "u":  "terofita / emicriptofita",
    "a":  "idrofita",
    "k":  "emicriptofita monocarpica",
    "k*": "biennale o poco vivace",
    "d":  "nanofanerofita-emicriptofita",
    "f":  "camefita-emicriptofita",
    "s":  "pleustofita",
}

LIFEFORM_EN = {
    "p":  "deciduous phanerophyte",
    "i":  "evergreen phanerophyte",
    "n":  "deciduous nanophanerophyte",
    "j":  "evergreen nanophanerophyte",
    "z":  "woody chamaephyte",
    "c":  "herbaceous chamaephyte",
    "e":  "epiphyte",
    "h":  "hemicryptophyte",
    "g":  "geophyte",
    "t":  "therophyte",
    "u":  "therophyte / hemicryptophyte",
    "a":  "hydrophyte",
    "k":  "monocarpic hemicryptophyte",
    "k*": "biennial or short-lived perennial",
    "d":  "nanophanerophyte-hemicryptophyte",
    "f":  "chamaephyte-hemicryptophyte",
    "s":  "pleustophyte",
}

LIFEFORM_DE = {
    "p":  "sommergrüner Phanerophyt",
    "i":  "immergrüner Phanerophyt",
    "n":  "sommergrüner Nanophanerophyt",
    "j":  "immergrüner Nanophanerophyt",
    "z":  "verholzter Chamaephyt",
    "c":  "krautiger Chamaephyt",
    "e":  "Epiphyt",
    "h":  "Hemikryptophyt",
    "g":  "Geophyt",
    "t":  "Therophyt",
    "u":  "Therophyt / Hemikryptophyt",
    "a":  "Hydrophyt",
    "k":  "monokarper Hemikryptophyt",
    "k*": "zwei- oder kurzlebig mehrjährig",
    "d":  "Nanophanerophyt-Hemikryptophyt",
    "f":  "Chamaephyt-Hemikryptophyt",
    "s":  "Pleustophyt",
}

LIFEFORM_FR = {
    "p":  "phanérophyte caduc",
    "i":  "phanérophyte sempervirent",
    "n":  "nanophanérophyte caduc",
    "j":  "nanophanérophyte sempervirent",
    "z":  "chaméphyte ligneux",
    "c":  "chaméphyte herbacé",
    "e":  "épiphyte",
    "h":  "hémicryptophyte",
    "g":  "géophyte",
    "t":  "thérophyte",
    "u":  "thérophyte / hémicryptophyte",
    "a":  "hydrophyte",
    "k":  "hémicryptophyte monocarpique",
    "k*": "bisannuelle ou vivace de courte durée",
    "d":  "nanophanérophyte-hémicryptophyte",
    "f":  "chaméphyte-hémicryptophyte",
    "s":  "pleustophyte",
}

LIFEFORM_BY_LANG = {"it": LIFEFORM_IT, "en": LIFEFORM_EN, "de": LIFEFORM_DE, "fr": LIFEFORM_FR}

# Flowering-chart month abbreviations, keyed the same as LANG.
MONTHS_BY_LANG = {
    "it": ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu", "Lug", "Ago", "Set", "Ott", "Nov", "Dic"],
    "en": ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
    "de": ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"],
    "fr": ["Janv", "Févr", "Mars", "Avr", "Mai", "Juin", "Juil", "Août", "Sept", "Oct", "Nov", "Déc"],
}

# Known single-letter codes for splitting combined forms without a separator
_LF_SINGLES = {"p","i","n","j","z","c","e","h","g","t","u","a","k","d","f","s"}


def _split_lifeform(code: str) -> list[str]:
    """Splits combined codes (e.g. 'n-p', 'ch', 'k*-t') into individual parts."""
    # First try with an explicit separator
    if "-" in code:
        return [p.strip() for p in code.split("-") if p.strip()]
    # Then try splitting character by character against known codes
    parts, i = [], 0
    while i < len(code):
        # Try k* first (2 chars)
        if code[i:i+2] in LIFEFORM_IT:
            parts.append(code[i:i+2]); i += 2
        elif code[i] in _LF_SINGLES:
            parts.append(code[i]); i += 1
        else:
            parts.append(code[i]); i += 1  # leave unchanged if unknown
    return parts


def tr_origin(code: str) -> str:
    return ORIGIN_BY_LANG[LANG].get((code or "").strip(), code or "")


def tr_lifeform(code: str) -> str:
    code = (code or "").strip()
    if not code:
        return ""
    # Strip 'w' prefix (casual/naturalised — does not change the life form)
    raw = code.lstrip("w") or code
    parts = _split_lifeform(raw)
    translated = [LIFEFORM_BY_LANG[LANG].get(p, p) for p in parts]
    val = " / ".join(dict.fromkeys(translated))  # deduplicate preserving order
    return val[:1].upper() + val[1:] if val else val


# ── TEXT HELPERS ──────────────────────────────────────────────────────────────

# Rank words for author removal
_RANKS = {"subsp.", "var.", "f.", "ssp.", "subvar.", "nothovar.", "nothosubsp."}


_AGGR_RANKS = {"aggr", "superaggr"}


def clean_taxon(name: str, rangstufe: str = "") -> str:
    """Removes the botanical author; appends '(aggr.)' for aggregates."""
    if not name:
        return ""
    words = name.split()
    if len(words) <= 2:
        result = name
    elif words[2] in _RANKS and len(words) > 3:
        result = " ".join(words[:4])
    else:
        result = " ".join(words[:2])
    if (rangstufe or "").strip() in _AGGR_RANKS:
        result += " (aggr.)"
    return result


def _baseline(mm_from_top: float, pt_size: float) -> float:
    """ReportLab Y (from bottom) for the baseline, with first-baseline = ascent.

    Ascent Helvetica Neue ≈ 0.78 × pt_size.
    """
    ascent_mm = 0.78 * pt_size * (25.4 / 72)
    return CH - (mm_from_top + ascent_mm) * mm


def _wrap_line(text: str, font: str, pt_size: float, max_w_pt: float) -> list:
    """Splits text into at most 2 lines that fit within max_w_pt (points). Returns a list.
    If no split fits within the limit, returns the split that minimises
    the widest line (visible overflow but complete text in correct order).
    """
    if not text:
        return [""]
    if stringWidth(text, font, pt_size) <= max_w_pt:
        return [text]
    words = text.split()
    if len(words) == 1:
        return [text]
    best, best_max = None, float("inf")
    fallback, fallback_max = None, float("inf")
    for i in range(1, len(words)):
        l1 = " ".join(words[:i])
        l2 = " ".join(words[i:])
        w1 = stringWidth(l1, font, pt_size)
        w2 = stringWidth(l2, font, pt_size)
        mx = max(w1, w2)
        if mx < fallback_max:
            fallback_max, fallback = mx, [l1, l2]
        if mx < best_max and w1 <= max_w_pt and w2 <= max_w_pt:
            best_max, best = mx, [l1, l2]
    return best or fallback or [text]


def _draw_name(c: canvas.Canvas, taxon: str, x: float, y_top_mm: float,
               max_w_pt: float, pt_max: float = 12, pt_min: float = 8):
    """Draws the taxonomic name in italic.
    Strategy: 1 line at pt_max → 2 lines at pt_max → 2 lines with font scaled down to pt_min.
    y_top_mm is the top edge of the first line (label coordinates, mm from top).
    """
    font = "HN-LightIt"
    lh_mm = pt_max * 1.2 * 25.4 / 72   # leading in mm (used only for the 2nd line)

    # 1 line at pt_max
    if stringWidth(taxon, font, pt_max) <= max_w_pt:
        c.setFont(font, pt_max)
        c.drawString(x, _baseline(y_top_mm, pt_max), taxon)
        return

    # 2 lines: find the best split, then scale if needed
    words = taxon.split()
    best_split, best_max_w = None, float("inf")
    for i in range(1, len(words)):
        l1 = " ".join(words[:i])
        l2 = " ".join(words[i:])
        mx = max(stringWidth(l1, font, pt_max), stringWidth(l2, font, pt_max))
        if mx < best_max_w:
            best_max_w, best_split = mx, (l1, l2)

    if best_split:
        l1, l2 = best_split
        # Scale the font until both lines fit
        pt = pt_max
        while pt > pt_min:
            if max(stringWidth(l1, font, pt), stringWidth(l2, font, pt)) <= max_w_pt:
                break
            pt -= 0.5
        lh = pt * 1.2 * 25.4 / 72
        c.setFont(font, pt)
        c.drawString(x, _baseline(y_top_mm, pt), l1)
        c.drawString(x, _baseline(y_top_mm + lh, pt), l2)
    else:
        # Single word: scale down to pt_min
        pt = pt_max
        while pt > pt_min and stringWidth(taxon, font, pt) > max_w_pt:
            pt -= 0.5
        c.setFont(font, pt)
        c.drawString(x, _baseline(y_top_mm, pt), taxon)


def _draw_stack_right(
    c: canvas.Canvas,
    text: str,
    font: str,
    pt_size: float,
    x_right: float,
    y_base: float,
    max_w_pt: float,
) -> float:
    """Draws text right-aligned; line 0 at y_base, subsequent lines going upward.
    Returns the y of the topmost line (for positioning the next block).
    """
    lh = pt_size * 1.2
    lines = _wrap_line(text, font, pt_size, max_w_pt)
    n = len(lines)
    c.setFont(font, pt_size)
    for i, line in enumerate(lines):
        # lines[0] is the first line (goes at the top → higher Y); lines[-1] at y_base
        c.drawRightString(x_right, y_base + (n - 1 - i) * lh, line)
    return y_base + (n - 1) * lh


# ── XLSX READING ──────────────────────────────────────────────────────────────

def read_xlsx(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get(COL_NAME):
            rows.append(d)
    return rows


def slug_from_file(file_val):
    if not file_val:
        return None
    return os.path.splitext(os.path.basename(str(file_val)))[0]


# ── DRAW LABEL ────────────────────────────────────────────────────────────────

def draw_label(c: canvas.Canvas, label: dict, ox: float, oy: float):
    """Draws the label back with origin (ox, oy) = bottom-left corner."""

    slug = slug_from_file(label.get("File"))

    # ── 1. FAMILY — HN Bold 12pt, left ───────────────────────────────────────
    # IDML: x=5mm y_top=16.2mm w=43.4mm h=4.7mm
    c.setFont("HN-Bold", 12)
    c.setFillColorRGB(0, 0, 0)
    c.drawString(
        ox + 5.0 * mm,
        oy + _baseline(16.2, 12),
        str(label.get(COL_FAMILY) or ""),
    )

    # ── 2. SPECIES NAME — HN Light Italic, left, wrap+scale if needed ─────────
    # IDML: x=5mm y_top=21.8mm w=43.4mm; space below down to y≈40mm
    taxon = clean_taxon(str(label.get(COL_NAME) or ""), str(label.get("Rangstufe") or ""))
    c.setFillColorRGB(0, 0, 0)
    _draw_name(c, taxon, ox + 5.0 * mm, 21.8, 58.2 * mm)

    # ── 3+4. LEBENSFORM + INDIGENAT — HN Light 6pt, right, max 2 lines ───────
    # Indigenat (lower field) anchored to its IDML baseline.
    # Lebensform is placed dynamically above the last line of Indigenat,
    # so the two blocks never overlap.
    c.setFillColorRGB(0.15, 0.15, 0.15)
    _MAX_W = 28.6 * mm          # right text column width
    _X_R   = ox + 63.2 * mm    # right edge
    _LH    = 6 * 1.2            # leading 7.2pt for 6pt text

    origin       = tr_origin(str(label.get(COL_ORIGIN) or ""))
    indig_base  = oy + _baseline(40.8, 6)
    indig_top   = _draw_stack_right(c, origin, "HN-Light", 6, _X_R, indig_base, _MAX_W)

    lifeform      = tr_lifeform(str(label.get(COL_LIFEFORM) or ""))
    lebens_base = indig_top + _LH
    _draw_stack_right(c, lifeform, "HN-Light", 6, _X_R, lebens_base, _MAX_W)

    # ── 5. IMAGES ─────────────────────────────────────────────────────────────
    def place_img(rel_path, ix, iy_top, iw, ih, aspect=True):
        if not rel_path:
            return
        p = str(rel_path)
        if not os.path.isabs(p):
            p = os.path.join(BASE_DIR, p)
        if not os.path.exists(p):
            return
        c.drawImage(
            p,
            ox + ix * mm,
            oy + CH - (iy_top + ih) * mm,
            width=iw * mm,
            height=ih * mm,
            preserveAspectRatio=aspect,
            anchor="sw",
        )

    # ── 5a/5b. ZEIGERWERTE + IUCN ─────────────────────────────────────────────
    _file_val = str(label.get("File") or "").strip()
    _slug     = os.path.splitext(os.path.basename(_file_val))[0] if _file_val else ""
    _zw_path  = os.path.join(OUTPUT_DIR, "zeigerwerte", f"{_slug}.jpg") if _slug else None
    _has_zw   = bool(_zw_path and os.path.exists(_zw_path))

    rl        = str(label.get(COL_IUCN) or "").strip().upper()
    _has_iucn = bool(rl and rl != "N/D" and rl in _IUCN_CATS)

    # base coordinates (full layout)
    _ZW_TOP_MM   = 44.0          # distance from label top to table top edge
    _ZW_H_MM     = 18.9          # zeigerwerte table height (normal layout)
    _IUCN_BOT_MM = 64.2 + 9.5   # distance from top to IUCN bottom edge = 73.7mm
    _IUCN_H_MM   = 9.5

    if _has_zw and _has_iucn:
        _draw_zeigerwerte(c, _zw_path,
            x_left=ox+5.0*mm, y_bottom=oy+CH-(_ZW_TOP_MM+_ZW_H_MM)*mm,
            total_w=58.2*mm,  h=_ZW_H_MM*mm)
        _draw_iucn_bar(c, rl,
            x_left=ox+5.0*mm, y_bottom=oy+CH-_IUCN_BOT_MM*mm,
            total_w=58.2*mm,  bar_h=_IUCN_H_MM*mm)
    elif _has_zw:
        _draw_zeigerwerte(c, _zw_path,
            x_left=ox+5.0*mm, y_bottom=oy+CH-(_ZW_TOP_MM+_ZW_H_MM)*mm,
            total_w=58.2*mm,  h=_ZW_H_MM*mm)
    elif _has_iucn:
        _draw_iucn_bar(c, rl,
            x_left=ox+5.0*mm, y_bottom=oy+CH-_IUCN_BOT_MM*mm,
            total_w=58.2*mm,  bar_h=_IUCN_H_MM*mm)
    # if neither: draw nothing

    # ── 5c. MAP (left) + FLOWERING (right) ───────────────────────────────────
    # Flowering: the 3 columns (1/6, 1/2, 5/6 of the image) centred below
    # IUCN circles 3,4,5 (VU,NT,LC). From c0=c3 and c2=c5: w=3/2*(c5-c3), ix=c3-w/6.
    _d_iucn  = (58.2 - 5 * _IUCN_GAP_MM) / 6
    _c3 = 5.0 + 3 * (_d_iucn + _IUCN_GAP_MM) + _d_iucn / 2
    _c5 = 5.0 + 5 * (_d_iucn + _IUCN_GAP_MM) + _d_iucn / 2
    _w_flow  = 1.5 * (_c5 - _c3)
    _ix_flow = _c3 - _w_flow / 6
    place_img(label.get(COL_MAP),  5.0,      75.0, 30.0,    19.0)
    place_img(_flowering_image_path(label), _ix_flow, 75.0, _w_flow, 19.0, aspect=False)

    # ── 6. LABEL BORDER ──────────────────────────────────────────────────────
    c.setStrokeColorRGB(0.75, 0.75, 0.75)
    c.setLineWidth(0.25)
    c.rect(ox, oy, CW, CH)




# ── A4 LAYOUT ─────────────────────────────────────────────────────────────────
# Label rotated 90°: each slot occupies CH × CW (99 × 68.2mm)
SLOT_W = CH          # slot width  = label height = 99mm
SLOT_H = CW          # slot height = label width  = 68.2mm
COLS   = 2
ROWS   = 4

PAGE_W, PAGE_H = A4  # 210 × 297mm in points (ReportLab)

MARGIN_X = (PAGE_W - COLS * SLOT_W) / (COLS + 1)   # ≈ 4.1mm
MARGIN_Y = (PAGE_H - ROWS * SLOT_H) / (ROWS + 1)   # ≈ 4.8mm

FOTO_MARGIN = 2.0   # mm of white margin around the photo on the front


def _slot_pos(page_pos: int, mirror_cols: bool = False):
    """Coordinates (ox, oy) of the bottom-left corner of the slot."""
    col = page_pos % COLS
    row = page_pos // COLS
    if mirror_cols:
        col = COLS - 1 - col
    ox = MARGIN_X + col * (SLOT_W + MARGIN_X)
    oy = PAGE_H - MARGIN_Y - (row + 1) * SLOT_H - row * MARGIN_Y
    return ox, oy


def _cut_lines(c: canvas.Canvas):
    """Crop marks as single lines at the page margins.
    One line per cut edge (no overlapping crosses).
    """
    c.saveState()
    c.setStrokeColorRGB(0.5, 0.5, 0.5)
    c.setLineWidth(0.25)
    TICK = 3.5 * mm   # tick length

    # X positions of columns: left and right edges of each column
    # + centre of internal gutters
    # Outer cut at MARGIN/2 from the page edge → same margin as the inner cut.
    # Inner cut at the centre of the gutter → MARGIN/2 for each adjacent label.
    x_cuts = []
    x_cuts.append(MARGIN_X / 2)                   # outer left: at half margin
    x_cuts.append(PAGE_W - MARGIN_X / 2)          # outer right: at half margin
    for col in range(1, COLS):
        ox_prev = MARGIN_X + (col - 1) * (SLOT_W + MARGIN_X)
        ox_curr = MARGIN_X + col * (SLOT_W + MARGIN_X)
        x_cuts.append((ox_prev + SLOT_W + ox_curr) / 2)   # centre of inner gutter

    y_cuts = []
    y_cuts.append(MARGIN_Y / 2)                   # outer bottom: at half margin
    y_cuts.append(PAGE_H - MARGIN_Y / 2)          # outer top: at half margin
    for row in range(1, ROWS):
        oy_bot_prev = PAGE_H - row * (MARGIN_Y + SLOT_H)
        y_cuts.append(oy_bot_prev - MARGIN_Y / 2)          # centre of inner gutter

    # Horizontal ticks on the left/right page edges (for Y cuts)
    for y in y_cuts:
        c.line(0,          y, TICK,          y)
        c.line(PAGE_W,     y, PAGE_W - TICK, y)

    # Vertical ticks on the top/bottom page edges (for X cuts)
    for x in x_cuts:
        c.line(x, PAGE_H,        x, PAGE_H - TICK)
        c.line(x, 0,             x, TICK)

    c.restoreState()


# ── PHOTO SELECTION ───────────────────────────────────────────────────────────
FOTO_DIR           = os.path.join(OUTPUT_DIR, "foto")
MAIN_PHOTOS        = os.path.join(OUTPUT_DIR, "images_ext")
MAIN_CANDIDATES    = os.path.join(OUTPUT_DIR, "images_ext_candidates")
SPECIAL_PHOTOS     = os.path.join(OUTPUT_DIR, "special_cases")
SPECIAL_CANDIDATES = os.path.join(OUTPUT_DIR, "special_cases_candidates")

_FETCH_H = {"User-Agent": "Mozilla/5.0"}
_FETCH_DELAY = 0.8


def _extract_original_url(imgproxy_url):
    try:
        path = urlparse(imgproxy_url).path
        b64  = path.split("/")[-1]
        b64  = re.sub(r"\.[a-z]+$", "", b64)
        padding = 4 - len(b64) % 4
        if padding != 4:
            b64 += "=" * padding
        return base64.urlsafe_b64decode(b64).decode("utf-8")
    except Exception:
        return imgproxy_url


def _clean_credit(credit):
    m = re.search(r'©\s*(?:Copyright\s*)?(.+)', credit)
    return ("© " + m.group(1).strip()) if m else credit.strip()


def _fetch_gallery(slug):
    """Returns a sorted list of (url, title) from the infoflora gallery, or []."""
    try:
        r = requests.get(
            f"https://www.infoflora.ch/en/flora/{slug}.html",
            headers=_FETCH_H, timeout=15,
        )
        if r.status_code != 200:
            return []
    except Exception:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    seen, results = set(), []
    for a in soup.select("#info-gallery a[title]"):
        href = a.get("href", "")
        orig = a.get("original-image") or (_extract_original_url(href) if href else "")
        if orig and "infoflora" in orig and orig not in seen:
            seen.add(orig)
            results.append((orig, a.get("title", "")))
    return results


def _subset_indices(n):
    """Replicates the sampling logic from phase1."""
    if n <= 4:
        return list(range(n))
    return sorted({0, n // 3, 2 * n // 3, n - 1})


def _credit_from_candidate(chosen_filename, gallery):
    """Maps the NN_ prefix to the exact credit of the chosen photo."""
    m = re.match(r'^(\d+)_', chosen_filename)
    if not m or not gallery:
        return ""
    file_idx = int(m.group(1)) - 1
    idxs = _subset_indices(len(gallery))
    if file_idx >= len(idxs):
        return ""
    return _clean_credit(gallery[idxs[file_idx]][1])


def _credit_priority(gallery):
    """Credit with standard priority: FH Lauber > FH > others."""
    if not gallery:
        return ""
    ranked = sorted(
        gallery,
        key=lambda x: (0 if "Konrad Lauber" in x[1] else 1)
                      if "Flora Helvetica" in x[1] else 2,
    )
    return _clean_credit(ranked[0][1])


def _find_chosen_candidate(slug):
    """Returns the filename of the first remaining candidate (regardless of its number),
    or None if no candidate directory exists or it is empty."""
    for base in (MAIN_CANDIDATES, SPECIAL_CANDIDATES):
        d = os.path.join(base, slug)
        if os.path.isdir(d):
            files = sorted(f for f in os.listdir(d) if f.lower().endswith(".jpg"))
            return files[0] if files else None
    return None


def _is_placeholder(path: str) -> bool:
    """True if the image is not a usable photo:
    - 'Photo not available' placeholder (800×600 entirely white)
    - botanical drawing or anatomical section (white background >70%)
    """
    try:
        img = Image.open(path)
        arr = np.array(img.convert("RGB"))
        if img.size == (800, 600) and float(arr.mean()) > 250:
            return True
        white = float(((arr > 230).all(axis=2)).mean())
        return white > 0.70
    except Exception:
        return False


def _has_real_photo(sp: dict) -> bool:
    """True if the species has a real photo (not a placeholder) in candidates or images_ext."""
    slug = slug_from_file(sp.get("File"))
    if not slug:
        return False
    src = _pick_candidate(slug)
    return bool(src) and not _is_placeholder(src)


def _pick_candidate(slug: str) -> str | None:
    """Looks for the photo for the given slug among candidates and direct downloads."""
    for cand_base in (MAIN_CANDIDATES, SPECIAL_CANDIDATES):
        cand_dir = os.path.join(cand_base, slug)
        if os.path.isdir(cand_dir):
            files = sorted(f for f in os.listdir(cand_dir) if f.lower().endswith(".jpg"))
            if files:
                return os.path.join(cand_dir, files[0])
    for photo_base in (MAIN_PHOTOS, SPECIAL_PHOTOS):
        p = os.path.join(photo_base, f"{slug}.jpg")
        if os.path.exists(p):
            return p
    return None


def update_credits_xlsx(xlsx_path: str, credit_updates: dict):
    """Writes updated credits to the xlsx. credit_updates: {taxonname: credit}"""
    if not credit_updates:
        return
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    name_col   = next(i for i, h in enumerate(headers) if h and "taxonname" in str(h).lower())
    credit_col = next(i for i, h in enumerate(headers) if h == "Credit")
    updated = 0
    for row in ws.iter_rows(min_row=2):
        name = str(row[name_col].value or "").strip()
        if name in credit_updates:
            row[credit_col].value = credit_updates[name]
            updated += 1
    wb.save(xlsx_path)
    print(f"  Credits updated in xlsx: {updated}")


def select_photo(label_list: list, xlsx_path: str | None = None):
    """
    Copies the selected photo to foto/{slug}.jpg.
    For species without a Credit and with a single candidate (or no candidates),
    re-fetches the exact credit and updates it in the in-memory species record and in the xlsx.
    """
    foto_dir = FOTO_DIR
    os.makedirs(foto_dir, exist_ok=True)

    print(f"\n{'─'*55}")
    print("PHASE A — Photo selection + credit update")
    print(f"{'─'*55}")

    credit_updates = {}

    for sp in label_list:
        slug = slug_from_file(sp.get("File"))
        if not slug:
            print(f"  ⚠  no slug for '{clean_taxon(str(sp.get(COL_NAME) or '?'))}'")
            continue

        # ── copy photo ────────────────────────────────────────────────────────
        dest = os.path.join(foto_dir, f"{slug}.jpg")
        if os.path.exists(dest):
            foto_tag = "↷ already present"
        else:
            src = _pick_candidate(slug)
            if src:
                shutil.copy2(src, dest)
                foto_tag = f"← {os.path.relpath(src, BASE_DIR)}"
            else:
                foto_tag = "✗ not found"

        # ── credit ────────────────────────────────────────────────────────────
        existing_credit = str(sp.get(COL_CREDIT) or "").strip()
        if existing_credit or str(sp.get("Status") or "").strip() in _PLACEHOLDER_STATUSES:
            print(f"  ✓  {slug}.jpg  {foto_tag}")
            continue

        chosen = _find_chosen_candidate(slug)
        gallery = _fetch_gallery(slug)
        time.sleep(_FETCH_DELAY)

        if chosen:
            credit = _credit_from_candidate(chosen, gallery)
            credit_src = f"candidate {chosen}"
        else:
            credit = _credit_priority(gallery)
            credit_src = "gallery"

        if credit:
            sp[COL_CREDIT] = credit
            credit_updates[str(sp.get(COL_NAME) or "").strip()] = credit
            print(f"  ✓  {slug}.jpg  {foto_tag}")
            print(f"     credit [{credit_src}]: {credit}")
        else:
            print(f"  ✓  {slug}.jpg  {foto_tag}")
            print(f"     credit: not found")

    if xlsx_path and credit_updates:
        print(f"\n  Persisting {len(credit_updates)} credits to {xlsx_path}…")
        update_credits_xlsx(xlsx_path, credit_updates)


# ── PRE-FLIGHT ────────────────────────────────────────────────────────────────

def preflight(label_list: list) -> bool:
    """Verifies all assets. Returns True if everything is OK."""
    print(f"\n{'─'*55}")
    print("PHASE B — Pre-flight check")
    print(f"{'─'*55}")

    all_ok = True

    for sp in label_list:
        slug  = slug_from_file(sp.get("File")) or "?"
        name  = clean_taxon(str(sp.get(COL_NAME) or ""))
        rl    = str(sp.get(COL_IUCN) or "").strip().upper()

        def _abs(p):
            if not p: return None
            p = str(p)
            return p if os.path.isabs(p) else os.path.join(BASE_DIR, p)

        checks = {
            "photo":     os.path.join(FOTO_DIR, f"{slug}.jpg"),
            "map":       _abs(sp.get(COL_MAP)),
            "flowering": _abs(sp.get(COL_FLOW)),
            "iucn":      os.path.join(IUCN_DIR, f"{rl}.jpg") if rl and rl != "N/D" else None,
        }

        missing = [k for k, v in checks.items() if v and not os.path.exists(v)]
        n_a     = [k for k, v in checks.items() if v is None]

        if missing:
            print(f"  ✗  {name} ({slug})")
            for k in missing:
                print(f"       missing: {k}")
            all_ok = False
        elif n_a:
            print(f"  △  {name}  [{', '.join(f'{k}=—' for k in n_a)}]")
        else:
            print(f"  ✓  {name}")

    verdict = "All assets present." if all_ok else "⚠ Missing assets — blank spaces in the PDF."
    print(f"\n  {verdict}")
    return all_ok


# ── PHOTO: EXIF + FILL/CROP ───────────────────────────────────────────────────

def _prepare_photo(src_path: str, target_w_pt: float, target_h_pt: float) -> str:
    """Opens the photo, corrects EXIF orientation, crops to fill the target (cover).
    Saves to a temporary file and returns its path.
    """
    img = Image.open(src_path)
    img = ImageOps.exif_transpose(img)      # correct EXIF rotation
    img = img.convert("RGB")

    iw, ih = img.size
    target_ratio = target_w_pt / target_h_pt
    img_ratio    = iw / ih

    if img_ratio > target_ratio:
        # image wider than target → scale by height, crop width
        new_w = round(ih * target_ratio)
        new_h = ih
    else:
        # image taller than target → scale by width, crop height
        new_w = iw
        new_h = round(iw / target_ratio)

    left = (iw - new_w) // 2
    top  = (ih - new_h) // 2
    img  = img.crop((left, top, left + new_w, top + new_h))

    tmp = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
    img.save(tmp.name, "JPEG", quality=95)
    tmp.close()
    return tmp.name


# ── PDF BACK ──────────────────────────────────────────────────────────────────

def generate_back(label_list: list, out_pdf: str):
    """A4 portrait, 2×4 labels rotated 90°, with crop marks."""
    print(f"\n{'─'*55}")
    print(f"PHASE C — Back  →  {out_pdf}")
    print(f"{'─'*55}")

    c = canvas.Canvas(out_pdf, pagesize=A4)
    per_page = COLS * ROWS

    for i, sp in enumerate(label_list):
        page_pos = i % per_page
        if page_pos == 0 and i > 0:
            _cut_lines(c)
            c.showPage()

        ox, oy = _slot_pos(page_pos)
        c.saveState()
        c.translate(ox + SLOT_W, oy)
        c.rotate(90)
        draw_label(c, sp, 0, 0)
        c.restoreState()

    _cut_lines(c)
    c.save()
    n_pages = (len(label_list) - 1) // per_page + 1
    print(f"  Saved: {out_pdf}  ({len(label_list)} labels, {n_pages} p.)")


# ── PDF FRONT ─────────────────────────────────────────────────────────────────

def _is_fh_credit(credit: str) -> bool:
    """True if the credit is Flora Helvetica (no need to show it on the front)."""
    return "Flora Helvetica" in str(credit or "")


def _credit_visibile(credit: str) -> str:
    """Returns the credit to display, or '' if FH or sentinel is unavailable."""
    c = str(credit or "").strip()
    if not c or c == "-" or _is_fh_credit(c):
        return ""
    return c


def _draw_photo(c: canvas.Canvas, sp: dict, ox: float, oy: float):
    """Photo rotated 90° with margin and, if not FH, credit in white."""
    slug = slug_from_file(sp.get("File"))
    if not slug:
        return

    foto_path = os.path.join(FOTO_DIR, f"{slug}.jpg")
    credit    = str(sp.get(COL_CREDIT) or "").strip()

    c.saveState()
    c.translate(ox + SLOT_W, oy)
    c.rotate(90)
    # Now in local label coordinates (CW × CH)

    m = FOTO_MARGIN * mm
    fw = CW - 2 * m
    fh = CH - 2 * m

    if os.path.exists(foto_path):
        tmp = _prepare_photo(foto_path, fw, fh)
        try:
            c.drawImage(tmp, m, m, width=fw, height=fh, preserveAspectRatio=False)
        finally:
            os.unlink(tmp)

    # Small white credit (only if not FH, not sentinel, and not empty)
    visible = _credit_visibile(credit)
    if visible:
        CREDIT_PT = 4.5
        c.setFont("HN-Light", CREDIT_PT)
        c.setFillColorRGB(1, 1, 1)
        c.drawRightString(m + fw - 1 * mm, m + 1.5 * mm, visible)

    c.restoreState()


def generate_front(label_list: list, out_pdf: str):
    """A4 portrait, 2×4 photos with margin; mirrored columns for long-edge duplex."""
    print(f"\n{'─'*55}")
    print(f"PHASE D — Front  →  {out_pdf}")
    print(f"{'─'*55}")

    c = canvas.Canvas(out_pdf, pagesize=A4)
    per_page = COLS * ROWS

    for i, sp in enumerate(label_list):
        page_pos = i % per_page
        if page_pos == 0 and i > 0:
            _cut_lines(c)
            c.showPage()

        ox, oy = _slot_pos(page_pos, mirror_cols=True)
        _draw_photo(c, sp, ox, oy)

    _cut_lines(c)
    c.save()
    n_pages = (len(label_list) - 1) // per_page + 1
    print(f"  Saved: {out_pdf}  ({len(label_list)} labels, {n_pages} p.)")


# ── COMBINED PDF ──────────────────────────────────────────────────────────────

def generate_combined(label_list: list, out_pdf: str):
    """Single PDF: for each batch, front (odd pages) + back (even pages).
    Long-edge duplex printing: each sheet already has front and back paired.
    """
    print(f"\n{'─'*55}")
    print(f"PHASE C — Combined  →  {out_pdf}")
    print(f"{'─'*55}")

    c = canvas.Canvas(out_pdf, pagesize=A4)
    per_page = COLS * ROWS
    n_pages  = (len(label_list) - 1) // per_page + 1

    for p in range(n_pages):
        batch = label_list[p * per_page : (p + 1) * per_page]

        # ── front page (mirrored columns for duplex) ─────────────────────────
        for i, sp in enumerate(batch):
            ox, oy = _slot_pos(i, mirror_cols=True)
            _draw_photo(c, sp, ox, oy)
        _cut_lines(c)
        c.showPage()

        # ── back page ─────────────────────────────────────────────────────────
        for i, sp in enumerate(batch):
            ox, oy = _slot_pos(i)
            c.saveState()
            c.translate(ox + SLOT_W, oy)
            c.rotate(90)
            draw_label(c, sp, 0, 0)
            c.restoreState()
        _cut_lines(c)
        if p < n_pages - 1:
            c.showPage()

    c.save()
    print(f"  Saved: {out_pdf}  ({len(label_list)} labels, {n_pages*2} p.)")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    global LANG

    args = sys.argv[1:]
    solo_foto = "--solo-foto" in args

    if "--lang" in args:
        i = args.index("--lang")
        if i + 1 >= len(args):
            print(f"Error: --lang requires a value ({'/'.join(LANG_CHOICES)})")
            sys.exit(1)
        LANG = args[i + 1].strip().lower()
        del args[i:i + 2]
        if LANG not in LANG_CHOICES:
            print(f"Error: unsupported --lang '{LANG}'. Choices: {', '.join(LANG_CHOICES)}")
            sys.exit(1)

    args = [a for a in args if not a.startswith("--")]

    if not args:
        print(__doc__)
        sys.exit(1)

    xlsx_path   = args[0]
    label_list = read_xlsx(xlsx_path)

    if not label_list:
        print("No species found.")
        sys.exit(0)

    base = os.path.splitext(xlsx_path)[0]

    if solo_foto:
        prima = len(label_list)
        label_list = [sp for sp in label_list if _has_real_photo(sp)]
        print(f"\n  --solo-foto: {len(label_list)}/{prima} species with a real photo")

    select_photo(label_list, xlsx_path=xlsx_path)

    preflight(label_list)
    generate_combined(label_list, base + "_stampa.pdf")

    print(f"\n{'═'*55}")
    print(f"  DONE")
    print(f"    {base}_stampa.pdf")
    print(f"{'═'*55}")


if __name__ == "__main__":
    main()
